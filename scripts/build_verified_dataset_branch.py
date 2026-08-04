#!/usr/bin/env python3
"""Build a Git branch containing only contact-verified baseball samples.

The branch is assembled from Git objects, so large audio/video blobs do not need
to be checked out or uploaded again. The current working tree is never staged or
modified by this script.
"""

from __future__ import annotations

import argparse
import csv
import io
import os
import re
import subprocess
import tempfile
from collections import Counter, defaultdict
from dataclasses import dataclass, field
from pathlib import Path

from openpyxl import load_workbook


REQUIRED_FILES = {"audio.wav", "video.mp4", "label.txt", "sample.csv", "source.txt"}
SAMPLE_PATH_RE = re.compile(
    r"^dataset/(?P<label>fly_ball|ground_ball)/(?P<collector>[^/]+)/(?P<sample_id>[^/]+)$"
)


@dataclass
class VerifiedSample:
    path: str
    sample_id: str
    label: str
    collector: str
    verification_source: str
    verification_detail: str
    contact_time: str = ""
    timing_was_corrected: bool = False
    original_event_start: str = ""
    original_event_end: str = ""
    final_event_start: str = ""
    final_event_end: str = ""
    evidence: dict[str, str] = field(default_factory=dict)


def run_git(repo: Path, args: list[str], *, input_data: bytes | None = None, env=None) -> bytes:
    command = ["git", *args]
    completed = subprocess.run(
        command,
        cwd=repo,
        input=input_data,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
        check=False,
        env=env,
    )
    if completed.returncode:
        raise RuntimeError(
            f"Git command failed ({' '.join(command)}):\n"
            f"{completed.stderr.decode('utf-8', errors='replace')}"
        )
    return completed.stdout


def normalize_id(value: object, prefix: str) -> str:
    match = re.search(r"(\d+)", str(value))
    if not match:
        raise ValueError(f"Cannot extract an ID from {value!r}")
    return f"{prefix}_{int(match.group(1)):04d}"


def add_sample(samples: dict[str, VerifiedSample], sample: VerifiedSample) -> None:
    existing = samples.get(sample.path)
    if existing:
        raise ValueError(
            f"Duplicate verified path {sample.path}: "
            f"{existing.verification_source} and {sample.verification_source}"
        )
    samples[sample.path] = sample


def read_binary_review(
    workbook_path: Path,
    *,
    label: str,
    prefix: str,
    samples: dict[str, VerifiedSample],
) -> None:
    workbook = load_workbook(workbook_path, read_only=True, data_only=True)
    sheet = workbook[workbook.sheetnames[0]]
    for row in sheet.iter_rows(values_only=True):
        if len(row) < 2 or row[1] != 1:
            continue
        sample_id = normalize_id(row[0], prefix)
        path = f"dataset/{label}/Codex_Workstation/{sample_id}"
        add_sample(
            samples,
            VerifiedSample(
                path=path,
                sample_id=sample_id,
                label=label,
                collector="Codex_Workstation",
                verification_source="human_binary_review",
                verification_detail=f"{workbook_path.name}: usable=1",
            ),
        )


def read_timing_review(csv_path: Path, samples: dict[str, VerifiedSample]) -> None:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            contact_time = row.get("correct_contact_time", "").strip()
            if not contact_time:
                continue
            sample_id = row["sample_id"].strip()
            path = f"dataset/fly_ball/Codex_Workstation/{sample_id}"
            corrected = row.get("original_time_correct", "").strip().upper() == "N"
            add_sample(
                samples,
                VerifiedSample(
                    path=path,
                    sample_id=sample_id,
                    label="fly_ball",
                    collector="Codex_Workstation",
                    verification_source="human_timing_review",
                    verification_detail=(
                        f"{csv_path.name}: original_time_correct="
                        f"{row.get('original_time_correct', '').strip().upper()}"
                    ),
                    contact_time=contact_time,
                    timing_was_corrected=corrected,
                    original_event_start=row.get("event_start", "").strip(),
                    original_event_end=row.get("event_end", "").strip(),
                ),
            )


def read_first_pass(csv_path: Path, samples: dict[str, VerifiedSample]) -> None:
    with csv_path.open("r", encoding="utf-8-sig", newline="") as handle:
        for row in csv.DictReader(handle):
            path = row["main_relative_path"].replace("\\", "/").strip("/")
            match = SAMPLE_PATH_RE.match(path)
            if not match:
                raise ValueError(f"Invalid dataset path in first-pass CSV: {path}")
            add_sample(
                samples,
                VerifiedSample(
                    path=path,
                    sample_id=match.group("sample_id"),
                    label=match.group("label"),
                    collector=match.group("collector"),
                    verification_source="local_first_pass_direct",
                    verification_detail=(
                        f"{row.get('task_batch', '')}; {row.get('review_provenance', '')}"
                    ),
                    contact_time=row.get("annotated_candidate_seconds", "").strip(),
                    evidence=row,
                ),
            )


def list_tree(repo: Path, ref: str) -> dict[str, tuple[str, str]]:
    output = run_git(repo, ["ls-tree", "-r", ref, "--", "dataset"]).decode("utf-8")
    entries: dict[str, tuple[str, str]] = {}
    for line in output.splitlines():
        metadata, path = line.split("\t", 1)
        mode, _kind, oid = metadata.split()
        entries[path] = (mode, oid)
    return entries


def read_blobs(repo: Path, oids: set[str]) -> dict[str, bytes]:
    process = subprocess.Popen(
        ["git", "cat-file", "--batch"],
        cwd=repo,
        stdin=subprocess.PIPE,
        stdout=subprocess.PIPE,
        stderr=subprocess.PIPE,
    )
    request = b"".join(f"{oid}\n".encode("ascii") for oid in sorted(oids))
    stdout, stderr = process.communicate(request)
    if process.returncode:
        raise RuntimeError(stderr.decode("utf-8", errors="replace"))
    blobs: dict[str, bytes] = {}
    stream = io.BytesIO(stdout)
    while True:
        header = stream.readline()
        if not header:
            break
        parts = header.decode("ascii").strip().split()
        if len(parts) != 3 or parts[1] != "blob":
            raise RuntimeError(f"Unexpected git cat-file header: {header!r}")
        oid, _kind, size_text = parts
        size = int(size_text)
        blobs[oid] = stream.read(size)
        if stream.read(1) != b"\n":
            raise RuntimeError(f"Malformed git cat-file response for {oid}")
    return blobs


def parse_sample_csv(content: bytes, path: str) -> tuple[list[str], dict[str, str]]:
    text = content.decode("utf-8-sig")
    reader = csv.DictReader(io.StringIO(text))
    rows = list(reader)
    if len(rows) != 1 or not reader.fieldnames:
        raise ValueError(f"{path}/sample.csv must contain exactly one data row")
    return list(reader.fieldnames), rows[0]


def corrected_sample_csv(
    fieldnames: list[str], row: dict[str, str], sample: VerifiedSample
) -> bytes:
    contact = float(sample.contact_time)
    start = max(0.0, contact - 0.05)
    end = contact + 0.05
    row = dict(row)
    row["event_start"] = f"{start:.6f}".rstrip("0").rstrip(".")
    row["event_end"] = f"{end:.6f}".rstrip("0").rstrip(".")
    sample.final_event_start = row["event_start"]
    sample.final_event_end = row["event_end"]
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=fieldnames, lineterminator="\n")
    writer.writeheader()
    writer.writerow(row)
    return output.getvalue().encode("utf-8")


def validate_and_prepare(
    repo: Path,
    tree: dict[str, tuple[str, str]],
    samples: dict[str, VerifiedSample],
) -> tuple[dict[str, bytes], list[str]]:
    replacements: dict[str, bytes] = {}
    warnings: list[str] = []
    sample_files: dict[str, dict[str, str]] = {}
    needed_oids: set[str] = set()
    for sample in samples.values():
        matching = {
            path.rsplit("/", 1)[-1]: path
            for path in tree
            if path.startswith(sample.path + "/")
        }
        missing = REQUIRED_FILES - set(matching)
        if missing:
            raise ValueError(f"{sample.path} is missing required files: {sorted(missing)}")
        sample_files[sample.path] = matching
        needed_oids.update(tree[matching[name]][1] for name in ("label.txt", "sample.csv", "source.txt"))
    blobs = read_blobs(repo, needed_oids)

    for sample in samples.values():
        matching = sample_files[sample.path]

        label_path = matching["label.txt"]
        label_text = blobs[tree[label_path][1]].decode("utf-8-sig").strip()
        if label_text != sample.label:
            raise ValueError(f"Label mismatch in {sample.path}: {label_text!r} != {sample.label!r}")

        sample_path = matching["sample.csv"]
        fieldnames, row = parse_sample_csv(blobs[tree[sample_path][1]], sample.path)
        if row.get("sample_id") != sample.sample_id or row.get("label") != sample.label:
            raise ValueError(f"sample.csv identity mismatch in {sample.path}")
        try:
            start = float(row["event_start"])
            end = float(row["event_end"])
        except (KeyError, TypeError, ValueError) as exc:
            raise ValueError(f"Invalid event interval in {sample.path}") from exc
        if not 0 <= start < end or end - start > 0.200001:
            raise ValueError(f"Invalid event interval in {sample.path}: {start}..{end}")

        sample.original_event_start = sample.original_event_start or row["event_start"]
        sample.original_event_end = sample.original_event_end or row["event_end"]
        sample.final_event_start = row["event_start"]
        sample.final_event_end = row["event_end"]
        if sample.timing_was_corrected:
            replacements[sample_path] = corrected_sample_csv(fieldnames, row, sample)

        source_path = matching["source.txt"]
        source_text = blobs[tree[source_path][1]].decode("utf-8-sig")
        if "video_title:" not in source_text or "video_url:" not in source_text:
            raise ValueError(f"Incomplete source traceability in {sample.path}")

        secondary_fields = (
            ["landing_zone", "strength", "trajectory_type"]
            if sample.label == "fly_ball"
            else ["region", "strength", "bounce"]
        )
        pending = [name for name in secondary_fields if row.get(name, "").strip().lower() in {"", "pending", "unknown"}]
        if pending:
            warnings.append(f"{sample.path}: secondary fields pending: {','.join(pending)}")
    return replacements, warnings


def manifest_bytes(samples: dict[str, VerifiedSample]) -> bytes:
    fields = [
        "dataset_path",
        "sample_id",
        "label",
        "collector",
        "verification_source",
        "verification_detail",
        "contact_time",
        "timing_was_corrected",
        "original_event_start",
        "original_event_end",
        "final_event_start",
        "final_event_end",
    ]
    output = io.StringIO(newline="")
    writer = csv.DictWriter(output, fieldnames=fields, lineterminator="\n")
    writer.writeheader()
    for sample in sorted(samples.values(), key=lambda item: item.path):
        writer.writerow(
            {
                "dataset_path": sample.path,
                "sample_id": sample.sample_id,
                "label": sample.label,
                "collector": sample.collector,
                "verification_source": sample.verification_source,
                "verification_detail": sample.verification_detail,
                "contact_time": sample.contact_time,
                "timing_was_corrected": "Y" if sample.timing_was_corrected else "N",
                "original_event_start": sample.original_event_start,
                "original_event_end": sample.original_event_end,
                "final_event_start": sample.final_event_start,
                "final_event_end": sample.final_event_end,
            }
        )
    return ("\ufeff" + output.getvalue()).encode("utf-8")


def report_bytes(
    samples: dict[str, VerifiedSample], warnings: list[str], base_ref: str
) -> bytes:
    by_label = Counter(sample.label for sample in samples.values())
    by_source = Counter(sample.verification_source for sample in samples.values())
    corrected = sum(sample.timing_was_corrected for sample in samples.values())
    text = f"""# Verified Dataset Branch

This branch is based on `{base_ref}` and its `dataset/` tree contains only samples accepted by either human review or the completed local first-pass contact review.

## Counts

- Total: {len(samples)}
- Fly ball: {by_label['fly_ball']}
- Ground ball: {by_label['ground_ball']}
- Human reviewed: {by_source['human_binary_review'] + by_source['human_timing_review']}
- Local first-pass direct: {by_source['local_first_pass_direct']}
- Human timing corrections materialized: {corrected}

## Inclusion Contract

- Every sample has `video.mp4`, `audio.wav`, `label.txt`, `sample.csv`, and `source.txt`.
- Folder, `label.txt`, and `sample.csv` class identities agree.
- Event intervals are valid and no longer than 0.2 seconds.
- Human rows marked usable are included; explicit incorrect timing rows are recentered to 0.1 seconds around the supplied contact time.
- Samples marked uncertain or needing repair in the local first pass are excluded.

## Scope

The verified decision concerns bat-ball contact presence, timing usability, and the binary `fly_ball` / `ground_ball` label needed for classifier experiments. Some secondary semantic fields may remain `pending`; see `SECONDARY_FIELD_WARNINGS.txt`.
"""
    return text.encode("utf-8")


def hash_blob(repo: Path, content: bytes) -> str:
    return run_git(repo, ["hash-object", "-w", "--stdin"], input_data=content).decode().strip()


def build_commit(
    repo: Path,
    *,
    base_ref: str,
    branch: str,
    tree: dict[str, tuple[str, str]],
    samples: dict[str, VerifiedSample],
    replacements: dict[str, bytes],
    generated: dict[str, bytes],
) -> str:
    git_dir = Path(run_git(repo, ["rev-parse", "--git-dir"]).decode().strip())
    if not git_dir.is_absolute():
        git_dir = repo / git_dir
    index_path = git_dir / "codex-verified-dataset.index"
    index_path.unlink(missing_ok=True)
    env = os.environ.copy()
    env["GIT_INDEX_FILE"] = str(index_path)
    try:
        run_git(repo, ["read-tree", base_ref], env=env)
        dataset_paths = [path for path in tree]
        removals = b"".join(f"0 {'0' * 40}\t{path}\n".encode("utf-8") for path in dataset_paths)
        run_git(repo, ["update-index", "--index-info"], input_data=removals, env=env)

        selected_entries: list[tuple[str, str, str]] = []
        for sample in samples.values():
            prefix = sample.path + "/"
            for path, (mode, oid) in tree.items():
                if not path.startswith(prefix):
                    continue
                if path in replacements:
                    oid = hash_blob(repo, replacements[path])
                selected_entries.append((mode, oid, path))
        for path, content in generated.items():
            selected_entries.append(("100644", hash_blob(repo, content), path))
        additions = b"".join(
            f"{mode} {oid}\t{path}\n".encode("utf-8")
            for mode, oid, path in sorted(selected_entries, key=lambda item: item[2])
        )
        run_git(repo, ["update-index", "--index-info"], input_data=additions, env=env)
        root_tree = run_git(repo, ["write-tree"], env=env).decode().strip()
        commit = run_git(
            repo,
            [
                "commit-tree",
                root_tree,
                "-p",
                base_ref,
                "-m",
                "Create contact-verified training dataset",
            ],
            env=env,
        ).decode().strip()
        run_git(repo, ["update-ref", f"refs/heads/{branch}", commit])
        return commit
    finally:
        index_path.unlink(missing_ok=True)


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--repo", type=Path, default=Path.cwd())
    parser.add_argument("--human-dir", type=Path, required=True)
    parser.add_argument(
        "--direct-pass",
        type=Path,
        default=Path("reports/local_contact_evidence_20260804/DIRECT_PASS_324.csv"),
    )
    parser.add_argument("--base-ref", default="origin/main")
    parser.add_argument("--branch", default="codex/contact-verified-binary-20260804")
    parser.add_argument("--create-branch", action="store_true")
    args = parser.parse_args()
    repo = args.repo.resolve()
    human_dir = args.human_dir.resolve()
    direct_pass = args.direct_pass if args.direct_pass.is_absolute() else repo / args.direct_pass

    samples: dict[str, VerifiedSample] = {}
    read_binary_review(human_dir / "1-212.xlsx", label="ground_ball", prefix="G", samples=samples)
    read_binary_review(
        human_dir / "205-408(1).xlsx", label="ground_ball", prefix="G", samples=samples
    )
    read_binary_review(human_dir / "305-507.xlsx", label="fly_ball", prefix="F", samples=samples)
    read_binary_review(human_dir / "508-710.xlsx", label="fly_ball", prefix="F", samples=samples)
    read_timing_review(human_dir / "F101_304.csv", samples)
    human_count = len(samples)
    read_first_pass(direct_pass, samples)

    tree = list_tree(repo, args.base_ref)
    replacements, warnings = validate_and_prepare(repo, tree, samples)
    report_root = "reports/verified_dataset_20260804"
    generated = {
        f"{report_root}/VERIFIED_DATASET_MANIFEST.csv": manifest_bytes(samples),
        f"{report_root}/README.md": report_bytes(samples, warnings, args.base_ref),
        f"{report_root}/SECONDARY_FIELD_WARNINGS.txt": (
            ("\n".join(warnings) + "\n") if warnings else "No secondary-field warnings.\n"
        ).encode("utf-8"),
        "scripts/build_verified_dataset_branch.py": Path(__file__).read_bytes(),
    }

    counts = Counter(sample.label for sample in samples.values())
    print(f"human_verified={human_count}")
    print(f"first_pass_direct={len(samples) - human_count}")
    print(f"total_verified={len(samples)}")
    print(f"fly_ball={counts['fly_ball']}")
    print(f"ground_ball={counts['ground_ball']}")
    print(f"timing_corrections={len(replacements)}")
    print(f"secondary_field_warnings={len(warnings)}")
    if not args.create_branch:
        print("dry_run=passed")
        return

    existing = subprocess.run(
        ["git", "show-ref", "--verify", "--quiet", f"refs/heads/{args.branch}"],
        cwd=repo,
        check=False,
    )
    if existing.returncode == 0:
        raise RuntimeError(f"Local branch already exists: {args.branch}")
    if existing.returncode not in {0, 1}:
        raise RuntimeError(f"Unable to check local branch: {args.branch}")
    commit = build_commit(
        repo,
        base_ref=args.base_ref,
        branch=args.branch,
        tree=tree,
        samples=samples,
        replacements=replacements,
        generated=generated,
    )
    print(f"branch={args.branch}")
    print(f"commit={commit}")


if __name__ == "__main__":
    main()
